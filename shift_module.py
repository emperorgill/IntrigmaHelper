# shift_module.py - Supplies the Shift class and associated helper functions.

# Minor awkwardness - a lookup table in the form of an XLSX file is needed so
# that we can create Shift objects from just the short shift name (e.g. "4m").
# The lookup table is a global variable dictionary that starts out empty.  If
# a function needs the lookup table, it will check to see if it's empty.  If it
# is, it'll call load_lookup_table to populate the table.  That way we only
# have to read the table from disk once.

# TO-DO
# 1.  convert_date() input checking
# 2.  More shift object property testing (AACC, RLS, EPRP, Call, etc)
# 3.  Fill in rest of table
# 4.  Shift.credit property (depends on day of week)

# Standard Library Imports
import datetime as dt
import pprint # for debugging

# Third Party Imports
import openpyxl as xl

#Global variables
LOOKUP_TABLE = {} # Will be loaded once by load_lookup_table and remain constant
                  # from then on

# Helper functions
def load_lookup_table(filename = "Config Files/SACROS Lookup Table.xlsx"):
    """Given a lookup table in the form of a properly formatted xlsx file,
    load the table into a global variable dictionary."""
    global LOOKUP_TABLE
    xlsx = xl.load_workbook(filename, data_only=True)
    sheet = xlsx.active
    # Each shift is a row in the xlsx file and the first field in a row is
    # the short name of a shift.  The remaining fields in a row are the
    # properties of that shift.  Create a dictionary with each key being the
    # short name of a shift and the value being a list containing the properties
    # of the shift.  Skip the first row because it just contains
    # human-readable column headers.
    for row in range(2, sheet.max_row+1):
        shift_name = None
        shift_properties = []
        for col in range(1, sheet.max_column+1):
            if (col == 1):
                shift_name = sheet.cell(row, col).value
            else:
                shift_properties.append(sheet.cell(row,col).value)
        LOOKUP_TABLE[shift_name] = shift_properties

def is_a_shift(shift_type):
    """Tell if the received shift_type (a string) is a shift type that we know
    about."""
    if not LOOKUP_TABLE:
        load_lookup_table()
    if (shift_type in LOOKUP_TABLE.keys()):
        return True
    else:
        return False

def convert_date(month, date, year):
    """Convert int month, date, year to a datetime object (likely for subsequent
    use in a call to the Shift constructor"""
    date_string = str(month) + "/" + str(date) + "/" + str(year)
    format_string = "%m/%d/%Y"
    return dt.datetime.strptime(date_string, format_string)

# Individual unit test functions
def test_shift_properties(shift, expected_properties):
    """Receive a Shift object and make sure its properties match those in the
    expected_properties list.  For unit testing, not production."""
    assert shift.short_name == expected_properties[0]
    assert shift.long_name == expected_properties[1]
    assert shift.location == expected_properties[2]
    assert shift.start_time == expected_properties[3]
    assert shift.end_time == expected_properties[4]
    assert shift.hours_paid == expected_properties[5]
    assert shift.is_PIT == expected_properties[6]
    return

# The actual Shift class
class Shift:
    def __init__(self, shift_name, date, doctor = None):
        if not LOOKUP_TABLE:
            load_lookup_table()
        self.short_name = shift_name #String
        self.long_name = LOOKUP_TABLE[shift_name][0] #String
        self.location = LOOKUP_TABLE[shift_name][1] #String
        self.start_time = date + dt.timedelta(hours=LOOKUP_TABLE[shift_name][2])
            #Datetime object
        # For end time, need to consider shifts that end past midnight.  End
        # time is a Datetime object.
        if (LOOKUP_TABLE[shift_name][3] >= LOOKUP_TABLE[shift_name][2]):
            self.end_time = date + dt.timedelta(hours=LOOKUP_TABLE[shift_name][3])
        else:
            self.end_time = date + dt.timedelta(hours=24) + \
                            dt.timedelta(hours=LOOKUP_TABLE[shift_name][3])
        self.hours_paid = dt.timedelta(hours=LOOKUP_TABLE[shift_name][4])
            #Timedelta object
        self.is_PIT = True if LOOKUP_TABLE[shift_name][5][0].lower == 'y' else False
            #Boolean
        self.doctor = doctor # String if a doctor is assigned, None otherwise

    def __str__(self):
        """Return a human-readable string representation of the Shift object"""
        output = self.short_name + " is a shift starting at " + \
            self.start_time.strftime("%H%M") + " on " + \
            self.start_time.strftime("%m/%d/%Y") + " for " + str(self.doctor)
        return output
        
# Unit test script for the Shift class
def unit_tests():
    print("shift_module.py supplies the Shift class and associated helper")
    print("functions.  When run directly (as you are doing now), it runs a")
    print("series of unit tests on the class and helper functions.\n")

    assert convert_date(12, 7, 1941) == dt.datetime(1941, 12, 7)
    print("convert_date() tests OK!\n")

    # Test shift with fractional start time    
    test_shift = Shift("630m", dt.datetime(1776, 7, 2))
    test_expected_properties = ["630m", "Sac 630a-3p", "Sacramento", 
                                dt.datetime(1776, 7, 2, 6, 30),
                                dt.datetime(1776, 7, 2, 15), 
                                dt.timedelta(hours=8.5), False]
    test_shift_properties(test_shift, test_expected_properties)
    print("0630m Shift object properties test OK!")
    # Test shift that crosses midnight
    test_shift = Shift("22M.", dt.datetime(1974, 8, 9))
    test_expected_properties = ["22M.", "Sac 10p-9a", "Sacramento", 
                                dt.datetime(1974, 8, 9, 22),
                                dt.datetime(1974, 8, 10, 9),
                                dt.timedelta(hours=11), False]
    print("22M. Shift object properties test OK!")
    # Test shift that starts at midnight
    test_shift = Shift("24r", dt.datetime(1963, 8, 28))
    test_expected_properties = ["24r", "Ros 12a-8a", "Roseville", 
                                dt.datetime(1963, 8, 28) + dt.timedelta(
                                    hours=23.99), dt.datetime(1963, 8, 29, 8),
                                    dt.timedelta(hours=8), False]
    # (dt.datetime + dt.timedelta construction above done because dt.datetime
    # can't take a float for hours but timedelta can
    print("24r Shift object propertiies test OK!")
    # Test shift that ends at midnight
    test_shift = Shift("16r", dt.datetime(1933, 3, 4))
    test_expected_properties = ["16r", "Ros 4p-12a", "Roseville",
                                dt.datetime(1933, 3, 4, 16),
                                dt.datetime(1933, 3, 5, 0),
                                dt.timedelta(hours=8), False]
    print("16r Shift object properties test OK!")
    # Test PIT shift (also with fractional end time)
    test_shift = Shift("8rp.", dt.datetime(1989, 11, 9), "Dr. Gorbachev")
    test_expected_properties = ["8rp.", "Ros 8a-3:30p", "Roseville",
                                dt.datetime(1989, 11, 9, 8),
                                dt.datetime(1989, 11, 9) + dt.timedelta(
                                    hours=15.5), dt.timedelta(hours=7.5), True]
    print("8rp. Shift object properties test OK!")

    print("\nHere's the __str__() for a Shift object...")
    print(test_shift.__str__() + "\n")

    # Clear LOOKUP_TABLE make sure is_valid_shift_type()'s as needed call to
    # load_lookup_table() is working well
    global LOOKUP_TABLE
    LOOKUP_TABLE = {}
    print("\nLOOKUP_TABLE has been cleared for further testing - " +
          str(LOOKUP_TABLE))

    # Now test is_valid_shift_type()
    assert is_a_shift("4r") == True
    print("\nis_a_shift() tests OK for true case!")
    assert is_a_shift("Gill da Great!") == False
    print("is_a_shift() tests OK for false case!")

if __name__ == "__main__":
    unit_tests()
